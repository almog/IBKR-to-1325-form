import os
import subprocess
import pandas as pd
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from parse_statement import parse_statement
from datetime import datetime

from statement_to_1325 import statement_to_1325_form
from parse_statement import parse_statement
from boi_currency_rate_download import download_usd_ils_for_current_tax_year
from columns import ColumnNames as Cols

def get_input_paths() -> tuple[str, str]:
    default_tax_year = str(datetime.now().year - 1)
    default_statement_path = f'{default_tax_year}_ibkr_report.csv'
    default_rate_csv_path = 'usd_ils.csv'

    statement_csv_path = input(f"Enter the IBKR statement CSV path (default: {default_statement_path}): ") or default_statement_path
    rate_csv_path = input(f"Enter the rate CSV file path (will try to find it at: {default_rate_csv_path} or otherwise fetch it from BOI): ") or default_rate_csv_path
    if not os.path.exists(rate_csv_path):
        download_usd_ils_for_current_tax_year()

    return statement_csv_path, rate_csv_path

def main():
    # TODO:
    # Find a way to break things into smaller logical unit,
    # preferrably without tricking oneself to believe that extracting things,
    # that depend on shared contract such as `result_df` is anything but piling layers of indirection.
    statement_csv_path, rate_csv_path = get_input_paths()

    schema_name_to_df = parse_statement(statement_csv_path)
    result_df = statement_to_1325_form(schema_name_to_df['Trades'], rate_csv_path)

    result_df.to_csv('1325_form.csv', index=False, float_format='%.2f', date_format='%d/%m/%Y')
    result_df.to_excel('1325_form.xlsx', sheet_name='1325', index=False, float_format='%.2f')

    # Reopen the file and format the date
    book = openpyxl.load_workbook('1325_form.xlsx')
    sheet = book.active

    # START STYLING
    def find_col_letter(sheet, col_name):
        col = next(cell for cell in sheet[1] if cell.value == col_name) # the 1 here refers to the first row
        return col.column_letter

    def set_style(sheet, cell_range, style):
        for cell in sheet[cell_range]:
            cell[0].style = style

    date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')
    ils_currency_style = NamedStyle(name='ils_currency', number_format = '₪#,##0.00')
    usd_currency_style = NamedStyle(name='usd_currency', number_format = '$#,##0.00')

    ils_currency_col_letters = [find_col_letter(sheet, col.value) for col in (Cols.REAL_LOSS_ILS,
                                                                          Cols.REAL_PROFIT_ILS,
                                                                          Cols.SELL_PRICE_ILS,
                                                                          Cols.BUY_PRICE_ILS)]
    for col_letter in ils_currency_col_letters:
        set_style(sheet, f'{col_letter}2:{col_letter}{sheet.max_row}', ils_currency_style)


    usd_col_letter = find_col_letter(sheet, Cols.SELL_PRICE_USD.value)
    set_style(sheet, f'{usd_col_letter}2:{usd_col_letter}{sheet.max_row}', usd_currency_style)
    sell_date_col_letter = find_col_letter(sheet, Cols.SELL_DATE.value)
    buy_date_col_letter = find_col_letter(sheet, Cols.BUY_DATE.value)
    set_style(sheet, f'{sell_date_col_letter}2:{sell_date_col_letter}{sheet.max_row}', date_style)
    set_style(sheet, f'{buy_date_col_letter}2:{buy_date_col_letter}{sheet.max_row}', date_style)
    # END STYLING


    # Adding row number column
    last_column = sheet.max_column + 1
    sheet.cell(row=1, column=last_column, value='מספר')  # set the header for the new column
    for i in range(2, sheet.max_row + 1):  # start from row 2 since row 1 is the header
        sheet.cell(row=i, column=last_column, value=i-1)  # write the row number, subtracting 1 to start from 1 instead of 2


    for column in sheet.columns:
        column = [cell for cell in column]
        col_width = max(len(str(cell.value)) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = col_width


    # Calculate the total sales, profit and loss table
    total_profit_loss = result_df[Cols.SELL_PRICE_USD.value].sum() - result_df[Cols.BUY_PRICE_USD.value].sum()
    print(f"Total Profit/Loss: {total_profit_loss}") # Not necessary for the 1325 form, but should match IBKR's statement

    total_sales_h1 = result_df[result_df[Cols.SELL_DATE.value].dt.month <= 6][Cols.SELL_PRICE_ILS.value].sum()
    total_sales_h2 = result_df[result_df[Cols.SELL_DATE.value].dt.month > 6][Cols.SELL_PRICE_ILS.value].sum()
    real_profit_h1 = result_df[result_df[Cols.SELL_DATE.value].dt.month <= 6][Cols.REAL_PROFIT_ILS.value].sum()
    real_profit_h2 = result_df[result_df[Cols.SELL_DATE.value].dt.month > 6][Cols.REAL_PROFIT_ILS.value].sum()
    real_loss_whole_year = result_df[Cols.REAL_LOSS_ILS.value].sum()

    # Add total sales, profit and loss
    additional_data = [
            ['', '', ''],
            ['', real_profit_h1, 'רווח מחצית ראשונה'],
            ['', real_profit_h2, 'רווח מחצית שנייה'],
            [real_loss_whole_year, '', 'סכום הפסד'],
            ['', total_sales_h1, 'סיכום מכירות מחצית ראשונה'],
            ['', total_sales_h2, 'סיכום מכירות מחצית שנייה'] ]

    for row_data in additional_data:
        sheet.append(row_data)
    # END Add total sales, profit and loss

    book.save('1325_form.xlsx')

    print("1325 form sales data saved")

if __name__ == "__main__":
    main()
